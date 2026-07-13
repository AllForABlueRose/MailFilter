"""HTTP-layer tests for Press: mailbox verification, compute, the Excel form, the
upload, and the two-press commit.

Outlook is stubbed throughout (there is no mock mode any more), so these cover both
worlds: with COM unreachable a mailbox check is *deferred* and every draft control
stays locked; with COM present the personal mailbox must BE the profile's address and
the shared one must be openable.
"""

import json
import shutil
import tempfile
import threading
import time
import unittest
from io import BytesIO
from pathlib import Path
from unittest import mock

import openpyxl

import config
from mailfilter import create_app, draft_ops, outlook, press
from tests.factories import make_mail
from tests.test_draft_ops import FakeNamespace, FakePythoncom, FakeReply

BODY = ("Dear {{ sender.first_name }},\n"
        "{% if row.uses_ftp %}Link: {{ ftp_link(row.file_name) }}"
        "{% else %}Attached: {{ row.file_name }}{% endif %}\n"
        "Ref {{ upper(row.ref) }}.")


class PressRouteTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.mkdtemp()
        tmp = Path(self._tmp)
        self._orig = {k: getattr(config, k) for k in (
            "CACHE_FILE", "SETTINGS_FILE", "TAGS_FILE", "TEMPLATES_DIR",
            "AUTOMATIONS_FILE", "CUSTOMERS_FILE", "COMPOSE_TEMPLATES_FILE",
            "PASSWORD_SETTINGS_FILE", "EXPERIMENTAL_FILE", "CUSTOMER_MATCH_FILE",
            "VAULT_FILE", "CALENDAR_PINS_FILE", "MAILBOX_FILE", "CATEGORIES_FILE", "WORKSPACE_DIR",
            "FILE_SERVER_DIR")}
        for key, name in [
                ("CACHE_FILE", "cache.json"), ("SETTINGS_FILE", "settings.json"),
                ("TAGS_FILE", "tags.json"), ("AUTOMATIONS_FILE", "auto.json"),
                ("CUSTOMERS_FILE", "cust.json"), ("COMPOSE_TEMPLATES_FILE", "ct.json"),
                ("PASSWORD_SETTINGS_FILE", "pwd.json"), ("EXPERIMENTAL_FILE", "exp.json"),
                ("CUSTOMER_MATCH_FILE", "cm.json"), ("VAULT_FILE", "vault.json"),
                ("CALENDAR_PINS_FILE", "cal.json"), ("MAILBOX_FILE", "mailbox.json"),
                ("CATEGORIES_FILE", "categories.json")]:
            setattr(config, key, tmp / name)
        config.TEMPLATES_DIR = tmp / "search_templates"
        config.WORKSPACE_DIR = tmp / "workspace"
        config.FILE_SERVER_DIR = tmp / "fileserver"
        config.FILE_SERVER_DIR.mkdir(parents=True, exist_ok=True)
        (config.FILE_SERVER_DIR / "inv.pdf").write_text("data")

        self.app = create_app()
        self.client = self.app.test_client()
        self.app.extensions["mail_store"].add_mails([
            make_mail(id="M1", conversation_id="C1", subject="Invoice",
                      sender="Alice Smith", sender_email="alice@acme.com",
                      received="2026-06-10 09:30:00",
                      recipient_emails=["me@example.com"], cc_emails=[],
                      attachments=[{"filename": "inv.pdf"}]),
            make_mail(id="M2", conversation_id="C2", subject="Drawings",
                      sender="Bob Lee", sender_email="bob@orion.com",
                      received="2026-06-11 10:00:00", attachments=[]),
        ])
        self.template = self.client.post("/api/compose-templates", json={
            "name": "Invoice", "body": BODY}).get_json()

    def tearDown(self):
        for k, v in self._orig.items():
            setattr(config, k, v)
        shutil.rmtree(self._tmp, ignore_errors=True)

    # ----- helpers -----

    def _items(self, row=None, template_id=True, mail_id="M1"):
        return [{"mail_id": mail_id,
                 "template_id": self.template["id"] if template_id else None,
                 "row": row or {}}]

    def _verify(self, kind="personal", address="me@example.com", profile=None):
        """Verify a mailbox against a stubbed Outlook profile."""
        with mock.patch.object(outlook, "is_available", return_value=True), \
             mock.patch.object(outlook, "profile_address",
                               return_value=profile or "me@example.com"), \
             mock.patch.object(outlook, "check_mailbox_access", return_value=True):
            return self.client.post("/api/press/mailbox",
                                    json={"kind": kind, "address": address}).get_json()

    def _mailboxes(self):
        return self.app.extensions["mailbox_store"]

    def _await_status(self, kind, status, timeout=5.0):
        """Wait for the off-thread verification to record its verdict."""
        deadline = time.monotonic() + timeout
        while time.monotonic() < deadline:
            box = self._mailboxes().get(kind)
            if box["status"] == status:
                return box
            time.sleep(0.01)
        return self._mailboxes().get(kind)

    # ----- mailbox verification -----

    def test_state_starts_unset_with_no_outlook(self):
        with mock.patch.object(outlook, "is_available", return_value=False):
            data = self.client.get("/api/press/state").get_json()
        self.assertFalse(data["outlook_available"])
        self.assertFalse(data["ready"])
        self.assertEqual(data["mailbox"]["personal"]["status"], "unset")
        self.assertTrue(any(f["id"] == "attachments" for f in data["filters"]))

    def test_a_mailbox_named_while_outlook_is_down_is_deferred_not_failed(self):
        with mock.patch.object(outlook, "is_available", return_value=False), \
             mock.patch.object(outlook, "profile_address",
                               side_effect=outlook.OutlookUnavailableError("no Outlook")):
            data = self.client.post("/api/press/mailbox", json={
                "kind": "personal", "address": "me@example.com"}).get_json()
        self.assertEqual(data["mailbox"]["status"], "pending")
        self.assertEqual(data["mailbox"]["address"], "me@example.com")  # remembered
        self.assertFalse(data["ready"])                                  # but locked

    def test_a_pending_check_completes_when_outlook_returns(self):
        with mock.patch.object(outlook, "is_available", return_value=False), \
             mock.patch.object(outlook, "profile_address",
                               side_effect=outlook.OutlookUnavailableError("no Outlook")):
            self.client.post("/api/press/mailbox",
                             json={"kind": "personal", "address": "me@example.com"})
        # Opening Press once Outlook is running retries the deferred check -- but off
        # the request thread, so the verdict lands just after the response rather than
        # in it. Entering a tab must never wait on a COM call that may not answer.
        with mock.patch.object(outlook, "is_available", return_value=True), \
             mock.patch.object(outlook, "profile_address", return_value="me@example.com"):
            self.client.get("/api/press/state")
            box = self._await_status("personal", "verified")
        self.assertEqual(box["status"], "verified")
        self.assertTrue(self.client.get("/api/press/state").get_json()["ready"])

    def test_opening_press_does_not_wait_on_a_slow_probe(self):
        """The deferred re-check must not hold the response open."""
        with mock.patch.object(outlook, "is_available", return_value=False), \
             mock.patch.object(outlook, "profile_address",
                               side_effect=outlook.OutlookUnavailableError("no Outlook")):
            self.client.post("/api/press/mailbox",
                             json={"kind": "personal", "address": "me@example.com"})

        released = threading.Event()

        def slow_probe():
            released.wait(10)
            return "me@example.com"

        with mock.patch.object(outlook, "is_available", return_value=True), \
             mock.patch.object(outlook, "profile_address", side_effect=slow_probe):
            started = time.monotonic()
            data = self.client.get("/api/press/state").get_json()
            elapsed = time.monotonic() - started
            self.assertLess(elapsed, 2)                            # returned; did not wait
            self.assertEqual(data["mailbox"]["personal"]["status"], "pending")

            # Let the probe finish and land its verdict before the temp cache is torn
            # down -- the point is that the *request* did not wait for it, not that the
            # check was abandoned.
            released.set()
            self.assertEqual(self._await_status("personal", "verified")["status"],
                             "verified")

    def test_the_personal_mailbox_must_be_the_profiles_own_address(self):
        data = self._verify(address="someone.else@example.com")
        self.assertEqual(data["mailbox"]["status"], "unset")
        self.assertEqual(data["mailbox"]["address"], "")     # dropped, so it is re-asked
        self.assertIn("not this Outlook profile's mailbox", data["mailbox"]["error"])
        self.assertFalse(data["ready"])

    def test_the_right_personal_address_verifies_case_insensitively(self):
        data = self._verify(address="ME@Example.com")
        self.assertEqual(data["mailbox"]["status"], "verified")
        self.assertTrue(data["ready"])

    def test_a_shared_mailbox_is_verified_by_opening_it(self):
        with mock.patch.object(outlook, "is_available", return_value=True), \
             mock.patch.object(outlook, "check_mailbox_access", return_value=True) as chk:
            data = self.client.post("/api/press/mailbox", json={
                "kind": "shared", "address": "team@example.com"}).get_json()
        chk.assert_called_once_with("team@example.com")
        self.assertEqual(data["mailbox"]["status"], "verified")

    def test_a_shared_mailbox_you_cannot_open_is_rejected(self):
        with mock.patch.object(outlook, "is_available", return_value=True), \
             mock.patch.object(outlook, "check_mailbox_access",
                               side_effect=outlook.OutlookUnavailableError("no access")):
            data = self.client.post("/api/press/mailbox", json={
                "kind": "shared", "address": "team@example.com"}).get_json()
        self.assertEqual(data["mailbox"]["status"], "unset")
        self.assertIn("no access", data["mailbox"]["error"])

    def test_settings_switch_the_mailbox_and_the_cc(self):
        self._verify()
        data = self.client.put("/api/press/settings",
                               json={"cc_enabled": False}).get_json()
        self.assertFalse(data["state"]["cc_enabled"])
        data = self.client.put("/api/press/settings",
                               json={"selected": "shared"}).get_json()
        self.assertFalse(data["ready"])   # shared is still unset

    def test_a_bad_kind_is_rejected(self):
        r = self.client.post("/api/press/mailbox", json={"kind": "nope", "address": "x"})
        self.assertEqual(r.status_code, 400)

    # ----- compute -----

    def test_an_item_with_no_template_is_empty(self):
        data = self.client.post("/api/press/compute", json={
            "items": self._items(template_id=False)}).get_json()
        self.assertEqual(data["results"][0]["status"], "empty")

    def test_a_missing_required_cell_fails_with_the_reason(self):
        data = self.client.post("/api/press/compute", json={
            "items": self._items({"file_name": "inv.pdf"})}).get_json()
        r = data["results"][0]
        self.assertEqual(r["status"], "failed")
        self.assertEqual(r["reasons"], ["missing row.ref"])
        self.assertEqual(data["columns"], ["uses_ftp", "file_name", "ref"])

    def test_filling_the_cell_makes_it_ok_and_carries_the_draft(self):
        data = self.client.post("/api/press/compute", json={
            "items": self._items({"file_name": "inv.pdf", "ref": "acme-1"})}).get_json()
        r = data["results"][0]
        self.assertEqual(r["status"], "ok")
        self.assertIn("Ref ACME-1.", r["plan"]["body"])
        self.assertTrue(r["plan"]["attachment"]["exists"])
        self.assertEqual(r["variables"], ["uses_ftp", "file_name", "ref"])

    def test_compute_writes_nothing(self):
        self.client.post("/api/press/compute", json={
            "items": self._items({"file_name": "inv.pdf", "ref": "a"})})
        self.assertFalse((config.WORKSPACE_DIR).exists())

    def test_too_many_items_are_refused(self):
        items = [{"mail_id": "M1", "template_id": None, "row": {}}
                 for _ in range(config.PRESS_MAX_ITEMS + 1)]
        r = self.client.post("/api/press/compute", json={"items": items})
        self.assertEqual(r.status_code, 400)

    # ----- the Excel form -----

    def _form(self, mail_ids=("M1", "M2"), rows=None):
        data = self.client.post("/api/press/form", json={
            "template_id": self.template["id"], "mail_ids": list(mail_ids),
            "rows": rows or {}}).get_json()
        return data, Path(data["folder"]) / data["name"]

    def test_the_form_lands_in_todays_workspace_with_the_right_columns(self):
        data, path = self._form()
        self.assertTrue(path.is_file())
        self.assertEqual(data["columns"],
                         [press.ENTRY_ID_COLUMN] + press.REPORT_COLUMNS
                         + ["uses_ftp", "file_name", "ref"])
        self.assertEqual(data["rows"], 2)

    def test_the_form_is_prefilled_from_the_mail(self):
        _data, path = self._form(("M1",))
        ws = openpyxl.load_workbook(path).active
        header = [c.value for c in ws[1]]
        row = dict(zip(header, [c.value for c in ws[2]]))
        self.assertEqual(row[press.ENTRY_ID_COLUMN], "M1")
        self.assertEqual(row["subject"], "Invoice")
        self.assertEqual(row["file_name"], "inv.pdf")
        # Blank for the user to fill in (openpyxl reads an empty cell back as None;
        # parse_xlsx normalizes both to "").
        self.assertFalse(row["ref"])

    def test_the_form_omits_the_entry_id_when_no_mail_is_loaded(self):
        data, _path = self._form(mail_ids=())
        self.assertNotIn(press.ENTRY_ID_COLUMN, data["columns"])
        self.assertEqual(data["rows"], 0)

    # ----- upload -----

    def _upload(self, path, mail_ids=("M1", "M2")):
        return self.client.post("/api/press/upload", data={
            "file": (BytesIO(path.read_bytes()), "form.xlsx"),
            "mail_ids": json.dumps(list(mail_ids)),
        }, content_type="multipart/form-data").get_json()

    def _fill(self, path, column, value):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col = header.index(column) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).value = value
        wb.save(path)

    def test_a_filled_form_binds_by_entry_id_and_makes_the_item_ok(self):
        _data, path = self._form(("M1",))
        self._fill(path, "ref", "acme-1042")
        up = self._upload(path, ("M1",))
        self.assertEqual(list(up["bound"]), ["M1"])
        self.assertEqual(up["unbound"], [])

        data = self.client.post("/api/press/compute", json={
            "items": self._items(up["bound"]["M1"])}).get_json()
        self.assertEqual(data["results"][0]["status"], "ok")
        self.assertIn("Ref ACME-1042.", data["results"][0]["plan"]["body"])

    def test_a_sheet_with_no_entry_id_binds_best_effort(self):
        # The form was downloaded with nothing loaded, so it has no Entry ID column.
        _data, path = self._form(mail_ids=())
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        ws.append(["" for _ in header])
        row = ws.max_row
        for col, value in [("subject", "Invoice"), ("datetime", "2026-06-10 09:30:00"),
                           ("sender", "alice@acme.com"), ("ref", "acme-9"),
                           ("file_name", "inv.pdf")]:
            ws.cell(row=row, column=header.index(col) + 1).value = value
        wb.save(path)

        up = self._upload(path)
        self.assertEqual(list(up["bound"]), ["M1"])
        self.assertEqual(up["bound"]["M1"]["ref"], "acme-9")

    def test_an_unmatchable_row_is_reported_not_guessed(self):
        _data, path = self._form(mail_ids=())
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        ws.append(["" for _ in header])
        ws.cell(row=ws.max_row, column=header.index("subject") + 1).value = "Nothing"
        wb.save(path)
        up = self._upload(path)
        self.assertEqual(up["bound"], {})
        self.assertIn("no loaded mail matches", up["unbound"][0]["reason"])

    def test_a_bad_upload_is_a_400(self):
        r = self.client.post("/api/press/upload", data={
            "file": (BytesIO(b"not a workbook"), "x.xlsx"), "mail_ids": "[]"},
            content_type="multipart/form-data")
        self.assertEqual(r.status_code, 400)

    # ----- create drafts -----

    def _commit(self, items, selected):
        reply = FakeReply()
        ns = FakeNamespace(reply)
        app = mock.Mock()
        app.GetNamespace.return_value = ns
        with mock.patch.object(draft_ops.outlook, "_import_pywin32",
                               return_value=(FakePythoncom, None, None)), \
             mock.patch.object(draft_ops.outlook, "_dispatch", return_value=app):
            resp = self.client.post("/api/press/create-drafts",
                                    json={"items": items, "selected": selected})
        return resp, reply

    def test_create_drafts_is_refused_until_the_mailbox_is_verified(self):
        resp = self.client.post("/api/press/create-drafts", json={
            "items": self._items({"file_name": "inv.pdf", "ref": "a"}),
            "selected": ["M1"]})
        self.assertEqual(resp.status_code, 409)
        self.assertIn("not been verified", resp.get_json()["description"])

    def test_a_verified_mailbox_drafts_and_never_sends(self):
        self._verify()
        items = self._items({"file_name": "inv.pdf", "ref": "acme-1"})
        resp, reply = self._commit(items, ["M1"])
        data = resp.get_json()
        self.assertEqual(data["created"], 1)
        self.assertTrue(reply.saved)
        self.assertFalse(reply.sent)
        self.assertEqual(reply.SentOnBehalfOfName, "me@example.com")
        self.assertEqual([r.address for r in reply.Recipients.added], ["me@example.com"])
        self.assertTrue(Path(data["audit"]).is_file())

    def test_the_cc_toggle_is_honoured(self):
        self._verify()
        self.client.put("/api/press/settings", json={"cc_enabled": False})
        items = self._items({"file_name": "inv.pdf", "ref": "a"})
        _resp, reply = self._commit(items, ["M1"])
        self.assertEqual(reply.Recipients.added, [])

    def test_an_item_the_server_computes_as_failed_is_never_drafted(self):
        # The client asks for it anyway; the server recomputes and refuses.
        self._verify()
        items = self._items({"file_name": "inv.pdf"})   # no ref -> failed
        resp, reply = self._commit(items, ["M1"])
        data = resp.get_json()
        self.assertEqual(data["created"], 0)
        self.assertEqual(data["requested"], 0)
        self.assertFalse(reply.saved)

    def test_an_unselected_item_is_not_drafted(self):
        self._verify()
        items = self._items({"file_name": "inv.pdf", "ref": "a"})
        resp, reply = self._commit(items, [])   # nothing ticked
        self.assertEqual(resp.get_json()["requested"], 0)
        self.assertFalse(reply.saved)

    def test_the_client_cannot_inject_draft_content(self):
        # A body smuggled into the request is ignored: the server recomputes the plan
        # from the stored template and the cache.
        self._verify()
        items = self._items({"file_name": "inv.pdf", "ref": "a"})
        items[0]["plan"] = {"body": "PWNED", "status": "ready", "mail_id": "M1"}
        _resp, reply = self._commit(items, ["M1"])
        self.assertNotIn("PWNED", reply.Body)
        self.assertIn("Ref A.", reply.Body)


class InternalDomainTests(unittest.TestCase):
    """`sender.is_internal` keys off the LAST SAVED identity, not Press's live mailbox.

    The mailbox is Press's business. The internal-domain set it feeds is everyone's --
    the mail list, org resolution and the SDS scan all classify senders with it. So an
    unset, pending or rejected mailbox must not silently empty it, which is what used
    to happen: the whole app changed behaviour because a Press check failed.
    """

    def setUp(self):
        self._tmp = tempfile.mkdtemp()
        tmp = Path(self._tmp)
        self._orig = {k: getattr(config, k) for k in (
            "CACHE_FILE", "SETTINGS_FILE", "TAGS_FILE", "TEMPLATES_DIR",
            "AUTOMATIONS_FILE", "CUSTOMERS_FILE", "COMPOSE_TEMPLATES_FILE",
            "PASSWORD_SETTINGS_FILE", "EXPERIMENTAL_FILE", "CUSTOMER_MATCH_FILE",
            "VAULT_FILE", "CALENDAR_PINS_FILE", "MAILBOX_FILE", "CATEGORIES_FILE",
            "WORKSPACE_DIR", "FILE_SERVER_DIR")}
        for key, name in [
                ("CACHE_FILE", "cache.json"), ("SETTINGS_FILE", "settings.json"),
                ("TAGS_FILE", "tags.json"), ("AUTOMATIONS_FILE", "auto.json"),
                ("CUSTOMERS_FILE", "cust.json"), ("COMPOSE_TEMPLATES_FILE", "ct.json"),
                ("PASSWORD_SETTINGS_FILE", "pwd.json"), ("EXPERIMENTAL_FILE", "exp.json"),
                ("CUSTOMER_MATCH_FILE", "cm.json"), ("VAULT_FILE", "vault.json"),
                ("CALENDAR_PINS_FILE", "cal.json"), ("MAILBOX_FILE", "mailbox.json"),
                ("CATEGORIES_FILE", "categories.json")]:
            setattr(config, key, tmp / name)
        config.TEMPLATES_DIR = tmp / "search_templates"
        config.WORKSPACE_DIR = tmp / "workspace"
        config.FILE_SERVER_DIR = tmp / "fileserver"
        config.FILE_SERVER_DIR.mkdir(parents=True, exist_ok=True)

        self.app = create_app()
        self.client = self.app.test_client()
        self.app.extensions["mail_store"].add_mails([
            make_mail(id="C1", conversation_id="X1", subject="Hi",
                      sender="Colleague", sender_email="colleague@example.com",
                      received="2026-06-10 09:30:00", attachments=[]),
        ])
        self.template = self.client.post("/api/compose-templates", json={
            "name": "Register",
            "body": '{{ if(sender.is_internal, "INTERNAL", "EXTERNAL") }}',
        }).get_json()

    def tearDown(self):
        for k, v in self._orig.items():
            setattr(config, k, v)
        shutil.rmtree(self._tmp, ignore_errors=True)

    def _register(self):
        """How the sender of the one cached mail is classified, right now."""
        data = self.client.post("/api/press/compute", json={"items": [
            {"mail_id": "C1", "template_id": self.template["id"], "row": {}}]}).get_json()
        return data["results"][0]["plan"]["body"].strip()

    def _verify_personal(self, address, profile=None):
        with mock.patch.object(outlook, "is_available", return_value=True), \
             mock.patch.object(outlook, "profile_address",
                               return_value=profile or address):
            return self.client.post("/api/press/mailbox", json={
                "kind": "personal", "address": address}).get_json()

    def _clear_personal(self):
        return self.client.post("/api/press/mailbox", json={
            "kind": "personal", "address": ""}).get_json()

    def test_the_sender_is_external_until_a_mailbox_is_proved(self):
        self.assertEqual(self._register(), "EXTERNAL")

    def test_a_verified_personal_mailbox_makes_its_domain_internal(self):
        self._verify_personal("me@example.com")
        self.assertEqual(self._register(), "INTERNAL")

    def test_clearing_the_mailbox_keeps_the_saved_internal_domain(self):
        self._verify_personal("me@example.com")
        box = self._clear_personal()
        self.assertEqual(box["mailbox"]["status"], "unset")   # Press is locked again...
        self.assertFalse(box["ready"])
        self.assertEqual(self._register(), "INTERNAL")        # ...but identity survives

    def test_a_rejected_mailbox_keeps_the_saved_internal_domain(self):
        self._verify_personal("me@example.com")
        # The user retypes a mailbox that is not this profile's: rejected, address
        # dropped. That is Press's problem and must not reclassify everyone's mail.
        data = self._verify_personal("someone@elsewhere.com", profile="me@example.com")
        self.assertEqual(data["mailbox"]["status"], "unset")
        self.assertEqual(self._register(), "INTERNAL")

    def test_a_deferred_check_keeps_the_saved_internal_domain(self):
        self._verify_personal("me@example.com")
        with mock.patch.object(outlook, "is_available", return_value=False), \
             mock.patch.object(outlook, "profile_address",
                               side_effect=outlook.OutlookUnavailableError("no Outlook")):
            data = self.client.post("/api/press/mailbox", json={
                "kind": "personal", "address": "me@example.com"}).get_json()
        self.assertEqual(data["mailbox"]["status"], "pending")
        self.assertEqual(self._register(), "INTERNAL")

    def test_proving_a_different_domain_moves_the_internal_domain(self):
        self._verify_personal("me@example.com")
        self.assertEqual(self._register(), "INTERNAL")
        self._verify_personal("me@newcorp.co.jp")
        # The identity moved, so the old colleague is now an outsider.
        self.assertEqual(self._register(), "EXTERNAL")
        self.assertEqual(
            self.app.extensions["mailbox_store"].own_address(), "me@newcorp.co.jp")

    def test_reproving_the_same_domain_leaves_the_saved_address_alone(self):
        self._verify_personal("me@example.com")
        self._verify_personal("me.again@example.com")
        self.assertEqual(
            self.app.extensions["mailbox_store"].own_address(), "me@example.com")
        self.assertEqual(self._register(), "INTERNAL")

    def test_a_shared_mailbox_never_touches_the_internal_domain(self):
        with mock.patch.object(outlook, "is_available", return_value=True), \
             mock.patch.object(outlook, "check_mailbox_access", return_value=True):
            self.client.post("/api/press/mailbox", json={
                "kind": "shared", "address": "team@partner.com"})
        self.assertEqual(self.app.extensions["mailbox_store"].own_address(), "")
        self.assertEqual(self._register(), "EXTERNAL")


if __name__ == "__main__":
    unittest.main()
