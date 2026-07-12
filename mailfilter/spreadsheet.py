"""Parse an uploaded Excel (.xlsx) sheet into Press rows.

The first sheet's first row is the header; every later row becomes a dict keyed
by the **normalized header** (trimmed, lowercased), plus stable aliases for the
fields Press reasons about (``config.BULK_COLUMNS`` -> ``row.subject``,
``row.file_name`` ...). Cells are coerced to strings so the rest of the pipeline
(matching, the template DSL) never has to think about Excel cell types; a
datetime cell is rendered with ``config.RECEIVED_FORMAT`` so it round-trips
against the cached mail's ``received`` string.

``openpyxl`` is imported lazily (mirroring the lazy pywin32 import in
``outlook.py``) so the rest of the app imports and runs without it installed.
Pure of Flask/COM.
"""

import logging
from datetime import date, datetime, time

import config

log = logging.getLogger(__name__)


class SpreadsheetError(ValueError):
    """An uploaded file that could not be read as an .xlsx sheet."""


def _normalize_header(value):
    return str(value or "").strip().lower()


def _cell_to_str(value):
    """Render one cell as the string the rest of the pipeline expects."""
    if value is None:
        return ""
    if isinstance(value, bool):  # before int: bool is an int subclass
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.strftime(config.RECEIVED_FORMAT)
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_xlsx(data, max_rows=None):
    """Parse ``.xlsx`` bytes into ``(headers, rows, dropped)``.

    ``headers`` is the list of normalized header names in column order. ``rows``
    is a list of dicts (one per non-empty data row). ``dropped`` is the number of
    rows beyond ``max_rows`` (default ``config.BULK_MAX_ROWS``) that were skipped.

    Raises :class:`SpreadsheetError` if the bytes are not a readable workbook or
    the sheet has no header row.
    """
    if max_rows is None:
        max_rows = config.BULK_MAX_ROWS
    try:
        from io import BytesIO

        import openpyxl
    except ImportError as e:
        raise SpreadsheetError(
            "Excel support unavailable (openpyxl not installed)"
        ) from e

    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise SpreadsheetError(f"could not read the file as .xlsx: {e}") from e
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise SpreadsheetError("the sheet is empty")

        headers = [_normalize_header(h) for h in header_row]
        if not any(headers):
            raise SpreadsheetError("the first row has no column headers")

        rows, dropped = [], 0
        for raw in rows_iter:
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue  # skip fully-blank rows
            if len(rows) >= max_rows:
                dropped += 1
                continue
            rows.append(_build_row(headers, raw))
        if dropped:
            log.warning("Bulk sheet exceeded %d rows; dropped %d", max_rows, dropped)
        return headers, rows, dropped
    finally:
        wb.close()


def _build_row(headers, raw):
    """One data row as ``{normalized_header: str}`` plus canonical aliases.

    Canonical aliases (``config.BULK_COLUMNS``) point at the same value as their
    source header, giving templates/matching stable names without losing access
    to the raw header. A blank header column is ignored.
    """
    row = {}
    for i, header in enumerate(headers):
        if not header:
            continue
        value = _cell_to_str(raw[i]) if i < len(raw) else ""
        row[header] = value
        alias = config.BULK_COLUMNS.get(header)
        # First mapped header wins an alias, so a sheet with both "date" and
        # "datetime" doesn't have them clobber each other.
        if alias and alias not in row:
            row[alias] = value
    return row


def write_xlsx(path, headers, rows):
    """Write a sheet of ``headers`` + ``rows`` (list of lists) to ``path``.

    The other half of ``parse_xlsx``: Press hands the user a **form** to fill in
    (the columns its chosen template actually needs) and reads the filled copy back
    through ``parse_xlsx``. Every value is written as text, so a reference like
    ``0012`` keeps its leading zeros and a datetime string round-trips against the
    cached ``received`` rather than being re-typed by Excel.

    ``openpyxl`` is imported lazily, exactly as in ``parse_xlsx``.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise SpreadsheetError(
            "Excel support unavailable (openpyxl not installed)"
        ) from e

    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        ws.title = "Press"
        ws.append([str(h) for h in headers])
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for row in rows:
            ws.append(["" if v is None else str(v) for v in row])
        # Force text so Excel doesn't reinterpret a reference or a datetime string.
        for column in ws.iter_cols(min_row=2):
            for cell in column:
                cell.number_format = "@"
        for i, header in enumerate(headers, start=1):
            width = max(12, min(42, len(str(header)) + 6))
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        wb.save(str(path))
    finally:
        wb.close()
    log.info("Wrote Press form (%d column(s), %d row(s)) to %s",
             len(headers), len(rows), path)
    return str(path)
